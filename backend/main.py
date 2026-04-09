"""
Hunter API — FastAPI backend.

Run with: python main.py
Server starts on http://localhost:8000
"""

import csv
import io
import json
from datetime import datetime

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from contextlib import asynccontextmanager

from database import init_db, get_connection
from models import (
    ScoreRequest, ScoreResponse, ScoreDetail, FirmResponse,
    FirmCreate, FirmUpdate, ScoreOverride,
    WEIGHTS, CRITERION_KEYS, compute_composite,
)
from scoring.engine import score_firm


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Initialize database on startup."""
    init_db()
    yield


app = FastAPI(
    title="Hunter by Trelity",
    description="AI-powered prospect evaluation and ranking",
    version="0.1.0",
    lifespan=lifespan,
)

# Allow Next.js dev server
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Helpers ─────────────────────────────────────────────────────────────────

# Column prefix mapping: criterion key → DB column prefix
PREFIX_MAP = {
    "cultural_alignment": "cultural",
    "growth_orientation": "growth",
    "industry_services": "industry",
    "revenue": "revenue",
    "employees": "employees",
    "geography": "geography",
}

# All score columns we SELECT (used in list_firms and get_firm)
SCORE_COLUMNS = """
    s.cultural_alignment, s.cultural_confidence, s.cultural_rationale, s.cultural_sources,
    s.cultural_override, s.cultural_override_note, s.cultural_override_at,
    s.growth_orientation, s.growth_confidence, s.growth_rationale, s.growth_sources,
    s.growth_override, s.growth_override_note, s.growth_override_at,
    s.industry_services, s.industry_confidence, s.industry_rationale, s.industry_sources,
    s.industry_override, s.industry_override_note, s.industry_override_at,
    s.revenue AS score_revenue, s.revenue_confidence, s.revenue_rationale, s.revenue_sources,
    s.revenue_override, s.revenue_override_note, s.revenue_override_at,
    s.employees AS score_employees, s.employees_confidence, s.employees_rationale, s.employees_sources,
    s.employees_override, s.employees_override_note, s.employees_override_at,
    s.geography, s.geography_confidence, s.geography_rationale, s.geography_sources,
    s.geography_override, s.geography_override_note, s.geography_override_at,
    s.composite, s.scored_at, s.score_notes, s.recommendation, s.is_real_score
"""


def _row_to_score_detail(row) -> ScoreDetail | None:
    """Build a ScoreDetail from a DB row, or None if no scores exist."""
    if row["composite"] is None:
        return None
    return ScoreDetail(
        cultural_alignment=row["cultural_alignment"],
        cultural_confidence=row["cultural_confidence"],
        cultural_rationale=row["cultural_rationale"],
        cultural_sources=row["cultural_sources"],
        cultural_override=row["cultural_override"],
        cultural_override_note=row["cultural_override_note"],
        cultural_override_at=row["cultural_override_at"],
        growth_orientation=row["growth_orientation"],
        growth_confidence=row["growth_confidence"],
        growth_rationale=row["growth_rationale"],
        growth_sources=row["growth_sources"],
        growth_override=row["growth_override"],
        growth_override_note=row["growth_override_note"],
        growth_override_at=row["growth_override_at"],
        industry_services=row["industry_services"],
        industry_confidence=row["industry_confidence"],
        industry_rationale=row["industry_rationale"],
        industry_sources=row["industry_sources"],
        industry_override=row["industry_override"],
        industry_override_note=row["industry_override_note"],
        industry_override_at=row["industry_override_at"],
        revenue=row["score_revenue"],
        revenue_confidence=row["revenue_confidence"],
        revenue_rationale=row["revenue_rationale"],
        revenue_sources=row["revenue_sources"],
        revenue_override=row["revenue_override"],
        revenue_override_note=row["revenue_override_note"],
        revenue_override_at=row["revenue_override_at"],
        employees=row["score_employees"],
        employees_confidence=row["employees_confidence"],
        employees_rationale=row["employees_rationale"],
        employees_sources=row["employees_sources"],
        employees_override=row["employees_override"],
        employees_override_note=row["employees_override_note"],
        employees_override_at=row["employees_override_at"],
        geography=row["geography"],
        geography_confidence=row["geography_confidence"],
        geography_rationale=row["geography_rationale"],
        geography_sources=row["geography_sources"],
        geography_override=row["geography_override"],
        geography_override_note=row["geography_override_note"],
        geography_override_at=row["geography_override_at"],
        composite=row["composite"],
        scored_at=row["scored_at"],
        score_notes=row["score_notes"],
        recommendation=row["recommendation"] if "recommendation" in row.keys() else None,
        is_real_score=row["is_real_score"] or 0,
    )


def _row_to_firm(row) -> FirmResponse:
    """Build a FirmResponse from a DB row."""
    return FirmResponse(
        id=row["id"],
        name=row["name"],
        tier=row["tier"],
        source=row["source"],
        website=row["website"] if "website" in row.keys() else None,
        city=row["city"],
        state=row["state"],
        employees=row["employees"],
        revenue_m=row["revenue_m"],
        bd_stage=row["bd_stage"],
        notes=row["notes"],
        last_contacted=row["last_contacted"],
        created_at=row["created_at"],
        updated_at=row["updated_at"],
        score=_row_to_score_detail(row),
    )


# ── POST /score ──────────────────────────────────────────────────────────────

@app.post("/score", response_model=ScoreResponse)
def score_endpoint(req: ScoreRequest):
    """
    Accept a firm name, score it, store results in DB, return scored JSON.
    If the firm already exists, re-scores and updates.
    """
    if req.criterion and req.criterion not in CRITERION_KEYS:
        raise HTTPException(status_code=400, detail=f"Invalid criterion: {req.criterion}. Valid: {CRITERION_KEYS}")

    scores = score_firm(req.name, criterion=req.criterion, refresh=req.refresh)
    conn = get_connection()
    cursor = conn.cursor()

    # Check if firm already exists
    cursor.execute("SELECT id FROM firms WHERE name = ?", (req.name,))
    row = cursor.fetchone()

    if row:
        firm_id = row["id"]
        # Update the existing score
        cursor.execute("DELETE FROM scores WHERE firm_id = ?", (firm_id,))
    else:
        # Insert new firm
        cursor.execute(
            "INSERT INTO firms (name) VALUES (?)",
            (req.name,),
        )
        firm_id = cursor.lastrowid

    # Insert score
    is_real = 0 if scores.get("score_notes", "").startswith("Stub") else 1
    cursor.execute(
        """INSERT INTO scores (
            firm_id,
            cultural_alignment, cultural_confidence, cultural_rationale, cultural_sources,
            growth_orientation, growth_confidence, growth_rationale, growth_sources,
            industry_services, industry_confidence, industry_rationale, industry_sources,
            revenue, revenue_confidence, revenue_rationale, revenue_sources,
            employees, employees_confidence, employees_rationale, employees_sources,
            geography, geography_confidence, geography_rationale, geography_sources,
            composite, score_notes, recommendation, is_real_score
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            firm_id,
            scores["cultural_alignment"], scores["cultural_confidence"],
            scores.get("cultural_rationale"), scores.get("cultural_sources"),
            scores["growth_orientation"], scores["growth_confidence"],
            scores.get("growth_rationale"), scores.get("growth_sources"),
            scores["industry_services"], scores["industry_confidence"],
            scores.get("industry_rationale"), scores.get("industry_sources"),
            scores["revenue"], scores["revenue_confidence"],
            scores.get("revenue_rationale"), scores.get("revenue_sources"),
            scores["employees"], scores["employees_confidence"],
            scores.get("employees_rationale"), scores.get("employees_sources"),
            scores["geography"], scores["geography_confidence"],
            scores.get("geography_rationale"), scores.get("geography_sources"),
            scores["composite"], scores["score_notes"], scores.get("recommendation"), is_real,
        ),
    )

    conn.commit()
    conn.close()

    return ScoreResponse(
        firm_id=firm_id,
        name=req.name,
        score=ScoreDetail(**scores),
    )


# ── GET /firms ───────────────────────────────────────────────────────────────

@app.get("/firms", response_model=list[FirmResponse])
def list_firms():
    """Return all firms with their latest scores, sorted by composite descending."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT f.*, {SCORE_COLUMNS}
        FROM firms f
        LEFT JOIN scores s ON s.firm_id = f.id
        ORDER BY s.composite DESC
    """)
    rows = cursor.fetchall()
    conn.close()
    return [_row_to_firm(row) for row in rows]


# ── GET /firms/{{id}} ──────────────────────────────────────────────────────────

@app.get("/firms/{firm_id}", response_model=FirmResponse)
def get_firm(firm_id: int):
    """Return a single firm with full score breakdown."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT f.*, {SCORE_COLUMNS}
        FROM firms f
        LEFT JOIN scores s ON s.firm_id = f.id
        WHERE f.id = ?
    """, (firm_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Firm not found")
    return _row_to_firm(row)


# ── POST /firms ──────────────────────────────────────────────────────────────

@app.post("/firms", response_model=FirmResponse)
def create_firm(firm: FirmCreate):
    """Add a new firm (single entry). Does not auto-score."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO firms (name, source, website) VALUES (?, ?, ?)",
        (firm.name, firm.source, firm.website),
    )
    firm_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return get_firm(firm_id)


# ── POST /firms/bulk ─────────────────────────────────────────────────────────

@app.post("/firms/bulk")
async def bulk_upload(file: UploadFile = File(...)):
    """
    Bulk import firms from CSV or XLSX.
    Expects columns: name (required), source (optional).
    Returns count of imported firms.
    """
    content = await file.read()
    firms_to_add = []

    if file.filename and file.filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            name_idx = next((i for i, h in enumerate(headers) if h in ("name", "firm name", "firm")), None)
            source_idx = next((i for i, h in enumerate(headers) if h in ("source", "source tag")), None)
            if name_idx is None:
                raise HTTPException(status_code=400, detail="No 'name' or 'firm name' column found")
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = str(row[name_idx]).strip() if row[name_idx] else ""
                source = str(row[source_idx]).strip() if source_idx is not None and row[source_idx] else None
                if name:
                    firms_to_add.append((name, source))
        except ImportError:
            raise HTTPException(status_code=400, detail="XLSX support requires openpyxl. Use CSV instead.")
    else:
        # CSV
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]
        name_col = next((f for f in reader.fieldnames or [] if f.strip().lower() in ("name", "firm name", "firm")), None)
        source_col = next((f for f in reader.fieldnames or [] if f.strip().lower() in ("source", "source tag")), None)
        if not name_col:
            raise HTTPException(status_code=400, detail="No 'name' or 'firm name' column found in CSV")
        for row in reader:
            name = (row.get(name_col) or "").strip()
            source = (row.get(source_col) or "").strip() if source_col else None
            if name:
                firms_to_add.append((name, source or None))

    if not firms_to_add:
        raise HTTPException(status_code=400, detail="No firms found in uploaded file")

    conn = get_connection()
    cursor = conn.cursor()
    added = 0
    skipped = 0
    for name, source in firms_to_add:
        cursor.execute("SELECT id FROM firms WHERE name = ?", (name,))
        if cursor.fetchone():
            skipped += 1
            continue
        cursor.execute("INSERT INTO firms (name, source) VALUES (?, ?)", (name, source))
        added += 1
    conn.commit()
    conn.close()
    return {"added": added, "skipped": skipped, "total": len(firms_to_add)}


# ── PATCH /firms/{{id}} ───────────────────────────────────────────────────────

@app.patch("/firms/{firm_id}", response_model=FirmResponse)
def update_firm(firm_id: int, update: FirmUpdate):
    """Update BD stage, last contacted, or append a timestamped note."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM firms WHERE id = ?", (firm_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Firm not found")

    if update.bd_stage is not None:
        cursor.execute(
            "UPDATE firms SET bd_stage = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (update.bd_stage, firm_id),
        )

    if update.last_contacted is not None:
        cursor.execute(
            "UPDATE firms SET last_contacted = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (update.last_contacted, firm_id),
        )

    if update.note_text is not None and update.note_text.strip():
        existing = row["notes"]
        try:
            notes_list = json.loads(existing) if existing else []
        except (json.JSONDecodeError, TypeError):
            notes_list = [{"ts": "", "text": existing}] if existing else []
        notes_list.insert(0, {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "text": update.note_text.strip(),
        })
        cursor.execute(
            "UPDATE firms SET notes = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (json.dumps(notes_list), firm_id),
        )

    conn.commit()
    conn.close()

    return get_firm(firm_id)


# ── PATCH /firms/{{id}}/scores/{{criterion}} ─────────────────────────────────

@app.patch("/firms/{firm_id}/scores/{criterion}", response_model=FirmResponse)
def override_score(firm_id: int, criterion: str, override: ScoreOverride):
    """
    Manual override for a single criterion score.
    Stores override value + note, recalculates composite using effective scores.
    """
    if criterion not in CRITERION_KEYS:
        raise HTTPException(status_code=400, detail=f"Invalid criterion: {criterion}")

    prefix = PREFIX_MAP[criterion]
    conn = get_connection()
    cursor = conn.cursor()

    # Verify firm and score exist
    cursor.execute("SELECT id FROM scores WHERE firm_id = ?", (firm_id,))
    score_row = cursor.fetchone()
    if not score_row:
        conn.close()
        raise HTTPException(status_code=404, detail="No scores found for this firm")

    # Set the override
    cursor.execute(f"""
        UPDATE scores SET
            {prefix}_override = ?,
            {prefix}_override_note = ?,
            {prefix}_override_at = CURRENT_TIMESTAMP
        WHERE firm_id = ?
    """, (override.score, override.note, firm_id))

    # Recalculate composite using effective scores (override if set, else AI score)
    cursor.execute("SELECT * FROM scores WHERE firm_id = ?", (firm_id,))
    row = cursor.fetchone()

    effective = {}
    for key in CRITERION_KEYS:
        p = PREFIX_MAP[key]
        ov = row[f"{p}_override"]
        ai = row[key]
        effective[key] = ov if ov is not None else ai

    new_composite = compute_composite(effective)
    cursor.execute(
        "UPDATE scores SET composite = ? WHERE firm_id = ?",
        (new_composite, firm_id),
    )

    conn.commit()
    conn.close()

    return get_firm(firm_id)


# ── GET /export ──────────────────────────────────────────────────────────────

@app.get("/export")
def export_csv():
    """Export the full ranked list as a CSV download."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT f.*, {SCORE_COLUMNS}
        FROM firms f
        LEFT JOIN scores s ON s.firm_id = f.id
        ORDER BY s.composite DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Firm Name", "Tier", "Source", "BD Stage", "Last Contacted", "Notes",
        "Growth Score", "Industry Score", "Revenue Score", "Culture Score",
        "Employees Score", "Geography Score", "Composite Score",
        "Growth Conf", "Industry Conf", "Revenue Conf", "Culture Conf",
        "Employees Conf", "Geography Conf",
    ])

    for row in rows:
        # Use effective scores (override if set, else AI)
        def eff(prefix, key):
            ov = row[f"{prefix}_override"]
            return ov if ov is not None else row[key]

        writer.writerow([
            row["name"],
            row["tier"],
            row["source"],
            row["bd_stage"],
            row["last_contacted"],
            row["notes"],
            eff("growth", "growth_orientation"),
            eff("industry", "industry_services"),
            eff("revenue", "score_revenue"),
            eff("cultural", "cultural_alignment"),
            eff("employees", "score_employees"),
            eff("geography", "geography"),
            row["composite"],
            row["growth_confidence"],
            row["industry_confidence"],
            row["revenue_confidence"],
            row["cultural_confidence"],
            row["employees_confidence"],
            row["geography_confidence"],
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=trelity_prospects.csv"},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
